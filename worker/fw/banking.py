"""Vendor bank details, decoupled from the Excel template.

In the desktop app the bank template .xlsx does double duty: its "Payment"
sheet defines the output columns *and* its "Mapping" sheet is the vendor
bank-detail database.  That works when the file lives on one person's
OneDrive.  It does not work for a hosted product -- every org needs its own
vendor list, and it needs to be queryable, auditable and editable by more
than one person.

So this module defines a `BankDetailsSource` interface with two
implementations:

  ExcelBankDetails  -- reads the existing Mapping sheet.  Lets an org onboard
                       by uploading the spreadsheet they already have.
  TableBankDetails  -- reads a list of records, i.e. what a `vendor_bank_details`
                       DB table returns.  This is where orgs end up.

The mapper only ever sees the interface, so the same workflow code runs
either way.  Onboarding path and steady state share one implementation.
"""

from __future__ import annotations

import re
from typing import Any, Protocol, runtime_checkable

_MAPPING_HINT = "mapping"


def norm_header(header: str) -> str:
    """Normalise a header to a lookup key: strip all whitespace, lowercase.

    "Sort Code" -> "sortcode", "Account Numb er" -> "accountnumber".
    Carried over from the desktop mapper so existing spreadsheets keep
    resolving to the same columns.
    """
    return re.sub(r"\s+", "", str(header)).lower()


def norm_vendor(name: str) -> str:
    return str(name).strip().lower()


def normalise_account(value: Any) -> str:
    """Zero-pad UK account numbers back to 8 digits (Excel eats leading zeros)."""
    s = _coerce_str(value)
    if not s:
        return ""
    return s.zfill(8) if s.isdigit() else s


def _coerce_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@runtime_checkable
class BankDetailsSource(Protocol):
    def field_keys(self) -> list[str]:
        """Normalised banking-field keys, in display order."""
        ...

    def lookup(self, vendor_name: str) -> dict[str, str]:
        """Banking fields for a vendor; empty dict when unknown."""
        ...


class TableBankDetails:
    """Bank details from records -- the shape a DB query returns.

    Each record is {"vendor": "Acme Ltd", "sortcode": "..", "accountnumber": ".."}.
    Keys are normalised on the way in so callers can pass raw column names.
    """

    def __init__(self, records: list[dict[str, Any]], vendor_key: str = "vendor"):
        self._keys: list[str] = []
        self._by_vendor: dict[str, dict[str, str]] = {}
        vkey = norm_header(vendor_key)

        for rec in records:
            normalised = {norm_header(k): v for k, v in rec.items()}
            vendor = normalised.pop(vkey, None)
            if not vendor:
                continue
            fields = {
                k: (normalise_account(v) if k == "accountnumber" else _coerce_str(v))
                for k, v in normalised.items()
            }
            for k in fields:
                if k not in self._keys:
                    self._keys.append(k)
            self._by_vendor[norm_vendor(vendor)] = fields

    def field_keys(self) -> list[str]:
        return list(self._keys)

    def lookup(self, vendor_name: str) -> dict[str, str]:
        return dict(self._by_vendor.get(norm_vendor(vendor_name), {}))


class ExcelBankDetails:
    """Bank details read from a bank template's Mapping sheet.

    Kept so an org can onboard with the spreadsheet they already maintain --
    and so this spike runs against Peter's real UK/US/EUR templates unchanged.
    """

    def __init__(self, template_path: str):
        from openpyxl import load_workbook

        wb = load_workbook(template_path, data_only=True, read_only=True)
        try:
            sheet = next(
                (s for s in wb.sheetnames if _MAPPING_HINT in s.lower()), None
            )
            if sheet is None:
                raise ValueError(
                    f"No mapping sheet found in {template_path!r}. "
                    f"Sheets present: {wb.sheetnames}"
                )
            ws = wb[sheet]

            col_meta: list[tuple[str, str]] = []
            self._by_vendor: dict[str, dict[str, str]] = {}

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    # Column 0 is always the vendor name, whatever it is titled.
                    for j, h in enumerate(row[1:]):
                        orig = str(h).strip() if h is not None else ""
                        col_meta.append(
                            (orig, norm_header(orig) if orig else f"_empty{j}")
                        )
                    continue

                vendor = row[0] if row else None
                if not vendor:
                    continue

                fields: dict[str, str] = {}
                for j, (_orig, key) in enumerate(col_meta):
                    raw = row[j + 1] if len(row) > j + 1 else None
                    fields[key] = (
                        normalise_account(raw)
                        if key == "accountnumber"
                        else _coerce_str(raw)
                    )
                self._by_vendor[norm_vendor(vendor)] = fields

            self._keys = [
                key for _orig, key in col_meta if not key.startswith("_empty")
            ]
        finally:
            wb.close()

    def field_keys(self) -> list[str]:
        return list(self._keys)

    def lookup(self, vendor_name: str) -> dict[str, str]:
        fields = self._by_vendor.get(norm_vendor(vendor_name))
        if fields is None:
            return {}
        # Blank header cells are placeholders, not banking fields — they must
        # not reach the review table or count towards "has bank details".
        return {k: fields.get(k, "") for k in self._keys}

    def as_records(self, vendor_key: str = "vendor") -> list[dict[str, Any]]:
        """Export for migrating a template into the vendor table on onboarding.

        Only real columns are exported. Mapping sheets carry trailing blank
        header cells (spare columns people left behind); those are placeholders,
        not banking fields, and must not become columns in the vendor table.
        """
        return [
            {vendor_key: vendor, **{k: fields.get(k, "") for k in self._keys}}
            for vendor, fields in self._by_vendor.items()
        ]


def payment_columns(template_path: str) -> list[str]:
    """Read the output column order from a bank template's payment sheet.

    Onboarding reads this once and stores it as the org's bank-file layout;
    after that no .xlsx is involved at run time.
    """
    from openpyxl import load_workbook

    wb = load_workbook(template_path, data_only=True, read_only=True)
    try:
        sheet = next(
            (s for s in wb.sheetnames if _MAPPING_HINT not in s.lower()),
            wb.sheetnames[0],
        )
        ws = wb[sheet]
        header_row = next(ws.iter_rows(values_only=True), ())
        return [str(h).strip() for h in header_row if h is not None and str(h).strip()]
    finally:
        wb.close()
